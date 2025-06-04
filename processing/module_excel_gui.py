from openpyxl import Workbook
from openpyxl.styles import Alignment
import math

def create_excel_worksheet():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Drop Data"
    return sheet, workbook

def create_headers(sheet):
    headers = [
        "Frame name", "Time [ms]",
        "Center X [pxl]", "Center Y [pxl]",
        "Horizontal Diameter Freefall [mm]",
        "Vertical Diameter Freefall [mm]",
        "Contact Width [mm]", "Contact Area [mm^2]",
        "Volume [mm^3]",
        "Left CA [⁰]", "Right CA [⁰]"
    ]

    sheet.append(headers)
    sheet.freeze_panes = "A2"

    for col_num, header in enumerate(headers, start=1):
        col_letter = sheet.cell(row=1, column=col_num).column_letter
        sheet.column_dimensions[col_letter].width = len(header) + 4
        sheet.cell(row=1, column=col_num).alignment = Alignment(horizontal="center", vertical="center")

def append_table_data(sheet, row):
    sheet.append([
        row[0],                     # Frame name
        row[1],                     # Time [ms]
        row[2], row[3],             # Center X/Y
        row[4], row[5],             # Hor/Vert diam
        row[6], row[7],             # Contact width, area
        row[8],                     # Volume
        row[9], row[10]            # Left/Right angle
    ])

def apply_alignment(sheet):
    align = Alignment(horizontal="center", vertical="center")
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = align

def set_up_format(sheet):
    format_list = [
        "@",    # Frame name
        "0.0000",    # Time
        "0", "0",    # Center X/Y
        "0.0000", "0.0000",  # Diameters
        "0.0000", "0.0000",  # Contact width & area
        "0.0000",            # Volume
        "0.00", "0.00"       # Contact angles
    ]

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for cell, fmt in zip(row, format_list):
            if cell.value is None:
                continue
            cell.number_format = fmt

def save_excel_file(workbook, path):
    workbook.save(path)  # может выбросить исключение, обрабатывается снаружи
